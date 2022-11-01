import requests
import json
import os
import discord
import asyncio
import re
from discord.ext import commands
from cogs.utils.common import CommonUtil
import iachara
import openpyxl

import time
import urllib.error
import urllib.request
from itertools import islice


def download_file(url, dst_path):
    try:
        with urllib.request.urlopen(url) as web_file:
            data = web_file.read()
            with open(dst_path, mode='wb') as local_file:
                local_file.write(data)
    except urllib.error.URLError as e:
        print(e)


def download_file_to_dir(url, dst_dir):
    download_file(url, os.path.join(dst_dir, os.path.basename(url)))


def dict_chunks(data, size):
    it = iter(data)
    for i in range(0, len(data), size):
        yield {k: data[k] for k in islice(it, size)}


class IacharaConfig(commands.Cog, name='iachara'):
    def __init__(self, bot):
        self.bot = bot
        self.c = CommonUtil()
        self.master_path = os.path.dirname(
            os.path.dirname(os.path.abspath(__file__)))

    @commands.command()
    async def excel(self, ctx, url):
        """いあきゃらAPIから取得したデータをエクセル出力する関数
        Args:
          :param url: いあきゃらのキャラシのURL
          :param ctx:
        """
        regex_url = "https://iachara.com/(sns|char)?/(?P<id>[0-9]+)/view"
        id = int(re.search(regex_url, url)["id"])
        transfer = iachara.Transfer(iachara.RequestsAdapter(id))
        name = transfer.get_name()
        status = transfer.get_status()
        skill = transfer.get_edit_skill()
        icon_url = transfer.get_icon_url()
        icon_filename = os.path.basename(icon_url)
        download_file_to_dir(icon_url, self.master_path + "/data/icon")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:C1")
        ws['A1'] = name
        ws.append(["", "", "", "STR", "CON", "POW", "DEX"])
        ws.append(["", "", "", status["str"], status["con"],
                  status["pow"], status["dex"]])
        ws.append(["", "", "", "APP", "SIZ", "INT", "EDU"])
        ws.append(["", "", "", status["app"], status["siz"],
                  status["int"], status["edu"]])
        ws.append(["", "", "", "HP", "MP", "SAN", "LUK"])
        ws.append(["", "", "", status["hp"], status["mp"],
                  status["san"], status["luck"]])

        ws.merge_cells("A2:C13")
        img = openpyxl.drawing.image.Image(
            self.master_path + "/data/icon/" + icon_filename)
        aspect = img.height / img.width
        img.width = 72 * 2.5
        img.height = img.width * aspect
        ws.add_image(img, 'A2')

        skill_groups = dict_chunks(skill, size=2)

        column = 9
        for skill_group in skill_groups:
            flag = 0
            for key, value in skill_group.items():
                if flag == 0:
                    ws["D" + str(column)] = key
                    ws["E" + str(column)] = value
                    flag = 1
                else:
                    ws["F" + str(column)] = key
                    ws["G" + str(column)] = value
                    column += 1

        wb.save(self.master_path + "/data/" + str(id) + ".xlsx")
        await ctx.reply(content="出力結果", file=discord.File(self.master_path + "/data/" + str(id) + ".xlsx"))
        os.remove(self.master_path + "/data/icon/" + icon_filename)
        os.remove(self.master_path + "/data/" + str(id) + ".xlsx")


async def setup(bot):
    await bot.add_cog(IacharaConfig(bot))
